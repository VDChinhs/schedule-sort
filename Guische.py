from tkinter import ttk,messagebox,filedialog,Menu,Tk
from openpyxl.styles import NamedStyle
import tkinter as tk
import openpyxl
import addlec
import addmon
import addnam
import manager as mn
import Guilec
import Guianalysis
import os
import readalpha as ra
import matplotlib.pyplot as plt
import numpy as np
import GoogleCalendar.main as ggcal

#GUI hiện thị lịch
class Guische:
    def __init__(self,window):
        
        self.cansave = False
        self.path = os.getcwd() + "\\alpha.xlsx"
        self.list_copy = []
        self.list_ = ""

        self.window = window
        self.window.title('Hỗ trợ xếp lịch giảng dạy')
        self.window.geometry('+200+150')
        self.window.iconbitmap("schedule.ico")
    
        self.frame = ttk.Frame(self.window)
        self.frame.pack()

        style = ttk.Style(self.window)

        self.window.tk.call("source", "theme-light.tcl")
        style.theme_use("theme-light")
        
        style.configure("My.TLabelframe.Label", font=("Helvetica", 13))
        style.configure("Custom.TButton", font=("Helvetica", 13))

        self.widgets_frame = ttk.LabelFrame(self.frame, text= "Xử lý",style="My.TLabelframe")
        self.widgets_frame.grid(row=0,column=0, padx=20, pady=10)

        self.name_entry = ttk.Entry(self.widgets_frame, font=("Helvetica", 14))
        # self.name_entry.insert(0,"Alpha")
        # self.name_entry.bind("<FocusIn>", lambda e: self.name_entry.delete('0', 'end'))
        self.name_entry.grid(row=1,column=0,sticky='ew')

        # self.button1 = ttk.Button(self.widgets_frame, text="Sửa", command=self.sua, style="Custom.TButton")
        # self.button1.grid(row=2, column=0, padx=5, pady=5, sticky="nsew")

        self.label1 = ttk.Label(self.widgets_frame, text="Kỳ đang chọn: ", font = ("Helvetica", 13))
        self.label1.grid(row=2, column=0, padx=5, pady=5, sticky="nsew")

        self.status_combobox = ttk.Combobox(self.widgets_frame, values=ra.listsheet(), font=("Helvetica", 14),state="readonly")
        if len(ra.listsheet()) != 0:
            self.status_combobox.current(0)
        self.status_combobox.grid(row=3, column=0, padx=5, pady=5,  sticky="ew")
        self.status_combobox.bind("<<ComboboxSelected>>", self.clearcombobox)
            
        self.separator = ttk.Separator(self.widgets_frame)
        self.separator.grid(row=4, column=0, padx=(20, 10), pady=10, sticky="ew")

        self.button1 = ttk.Button(self.widgets_frame, text="Sắp xếp cán bộ giảng dạy (Đầy đủ)", command=self.xepgiangvien, style="Custom.TButton")
        self.button1.grid(row=5, column=0, padx=5, pady=5, sticky="nsew")

        self.button1 = ttk.Button(self.widgets_frame, text="Sắp xếp cán bộ giảng dạy (Alpha)", command=self.xepgiangvien1, style="Custom.TButton")
        self.button1.grid(row=6, column=0, padx=5, pady=5, sticky="nsew")

        self.button1 = ttk.Button(self.widgets_frame, text="Kiểm tra lịch trùng", command=self.checkloptrung, style="Custom.TButton")
        self.button1.grid(row=7, column=0, padx=5, pady=5, sticky="nsew")

        self.button1 = ttk.Button(self.widgets_frame, text="Xuất File Excel", command=self.save, style="Custom.TButton")
        self.button1.grid(row=8, column=0, padx=5, pady=5, sticky="nsew")

        self.treeFrame = ttk.Frame(self.frame)
        self.treeFrame.grid(row=0, column=1, pady=10)
        self.treeScroll = ttk.Scrollbar(self.treeFrame)
        self.treeScroll.pack(side="right", fill="y")

        self.cols = ["STT","Mã học phần","Tên môn học", "Mã lớp học phần", "Thứ","Tiết","Tuần","Cán bộ giảng dạy","Lớp ghép","Trùng lịch"]
        self.treeview = ttk.Treeview(self.treeFrame, show="headings",yscrollcommand=self.treeScroll.set, columns=self.cols, height=35)
        for i in self.cols:
            if i == "STT":
                self.treeview.column(i, width=50,anchor= "center")
            elif i == "Tên môn học":
                self.treeview.column(i, width=330)
            elif i == "Mã lớp học phần":
                self.treeview.column(i, width=150,anchor= "center")
            elif i == "Thứ":
                self.treeview.column(i, width=50,anchor= "center")
            elif i == "Tiết":
                self.treeview.column(i, width=75,anchor= "center")
            elif i == "Tuần":
                self.treeview.column(i, width=210)
            elif i == "Cán bộ giảng dạy":
                self.treeview.column(i, width=150,anchor="center")
            elif i == "Lớp ghép":
                self.treeview.column(i, width=75,anchor= "center")
            elif i == "Trùng lịch":
                self.treeview.column(i, width=120,anchor= "center")
            else:
                self.treeview.column(i, width=100,anchor= "center")

        self.treeview.pack()
        self.treeScroll.config(command=self.treeview.yview) 

        style = ttk.Style()
        style.configure("Treeview.Heading", font=("Helvetica", 12))
        self.treeview.tag_configure("my_font", font=("Helvetica", 12))

        def on_treeview_cell_select(event):
            region = self.treeview.identify_region(event.x, event.y)
            selected_item = self.treeview.selection()
            if region == "cell":
                selected_column = self.treeview.identify_column(event.x)  
                selected_value = self.treeview.item(selected_item[0], "values")[int(selected_column[1:]) - 1]
                self.name_entry.delete('0','end')
                self.name_entry.insert(0,selected_value)

        self.treeview.bind("<ButtonRelease-1>", on_treeview_cell_select)

        # self.load_data()

        self.menubar = Menu(window)
        self.window.config(menu = self.menubar)

        self.fileMenu = Menu(self.menubar, tearoff = 0,font = ("Helvetica",13))
        self.menubar.add_cascade(label = "File", menu = self.fileMenu)
        self.fileMenu.add_cascade(label = "Mở File Excel",command= self.open)
        # self.fileMenu.add_cascade(label = "Edit",command= self.edit)
        self.fileMenu.add_cascade(label = "Đóng chương trình",command= self.close)

        self.window.bind('<Control-o>', lambda event: self.open())

        self.lecturerMenu = Menu(self.menubar, tearoff = 0,font = ("Helvetica",13))
        self.menubar.add_cascade(label = "Cán bộ giảng dạy", menu = self.lecturerMenu)
        self.lecturerMenu.add_cascade(label = "Xem thông tin", command= self.thongtin)
        self.lecturerMenu.add_cascade(label = "Thêm kỳ học", command= self.addnamhoc)
        self.lecturerMenu.add_cascade(label = "Thêm cán bộ giảng dạy", command= self.addgiangvien)
        self.lecturerMenu.add_cascade(label = "Thêm môn học", command= self.addmonhoc)

        self.thongke = Menu(self.menubar, tearoff = 0,font = ("Helvetica",13))
        self.menubar.add_cascade(label = "Phân tích thống kê", menu = self.thongke)
        self.thongke.add_cascade(label = "Danh sách lịch trùng", command= self.phantich)
        self.thongke.add_cascade(label = "Biểu đồ khối lượng giảng dạy", command= self.bieudo)

        self.taolich = Menu(self.menubar, tearoff = 0,font = ("Helvetica",13))
        self.menubar.add_cascade(label = "Tạo lịch Google Calendar",command=self.taolichgg)

    def close(self):
        self.window.quit()

    def open(self):
        self.file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx *.xls",)])
        if self.file_path:
            print("Thư mục được chọn:", self.file_path)
            self.treeview.delete(*self.treeview.get_children())

            try:
                if os.path.splitext(self.file_path)[1] == ".xlsx":
                    self.list_ = mn.readfile(self.file_path, readlec = True) 
                    
                if os.path.splitext(self.file_path)[1] == ".xls":
                    self.list_ = mn.readfilexls(self.file_path, readlec = True)
            except:
                messagebox.showwarning(title="Chú ý",message="File Excel không đúng định dạng")
                return

            for i in self.list_:
                self.list_copy.append(i)

            mn.checklopghep(self.list_)
            mn.check2lich(self.list_)
            self.load_data(self.list_)
            self.cansave = True
        else:
            print("Không có thư mục nào được chọn.")

    def edit(self):
        self.treeview.delete(*self.treeview.get_children())

    def thongtin(self):
        selected_table = self.status_combobox.get()

        if selected_table == "":
            messagebox.showwarning(title="Chú ý",message="Vui chọn kỳ học")
            return
        
        if not os.path.exists(self.path):
            messagebox.showwarning(title="Chú ý",message="Vui lòng thêm kỳ học")
            return
        if len(self.colfirst()) == 0 and len(self.rowfirst()) == 0:
            messagebox.showwarning(title="Chú ý",message="Không có thông tin cán bộ giảng dạy vui lòng thêm cán bộ giảng dạy và môn học")
        elif len(self.colfirst()) == 0 :
            messagebox.showwarning(title="Chú ý",message="Không có thông tin môn học vui lòng thêm môn học")
        elif len(self.rowfirst()) == 0 :
            messagebox.showwarning(title="Chú ý",message="Không có thông tin cán bộ giảng dạy vui lòng thêm cán bộ giảng dạy")
        else:
            Guilec.Guigiangvien(Tk(),selected_table,Guische,self.updatecombo)

    def clearcombobox(self,event):
            self.status_combobox.select_clear()

    def updatecombo(self):
        self.status_combobox["value"] = ra.listsheet()

    def addnamhoc(self):
        selected_table = self.status_combobox.get()
        self.second_window = tk.Toplevel(self.window)
        addnam.Addnam(self.second_window,self.window,Guische,selected_table,self.updatecombo,self.updatecombo)

    def addgiangvien(self):
        if self.status_combobox.get() == "":
            messagebox.showwarning(title="Chú ý",message="Vui chọn kỳ học")
            return
        
        if os.path.exists(self.path):
            selected_table = self.status_combobox.get()
            root = Tk()
            addlec.Addlec(root,self.window,Guische,selected_table)
        else:
            messagebox.showwarning(title="Chú ý",message="Vui lòng thêm kỳ học")

    def addmonhoc(self):
        if self.status_combobox.get() == "":
            messagebox.showwarning(title="Chú ý",message="Vui chọn kỳ học")
            return
        
        if os.path.exists(self.path):
            selected_table = self.status_combobox.get()
            root = Tk()
            addmon.Addmonhoc(root,self.window,Guische,selected_table)
        else:
            messagebox.showwarning(title="Chú ý",message="Vui lòng thêm kỳ học")
    def sua(self):
        pass

    def phantich(self):
        if self.cansave:
            if len(mn.ds_giangvien_trung(self.list_)) != 0:
                Guianalysis.Guianalysis(Tk(),self.list_)
            else:
                messagebox.showwarning(title="Chú ý",message="Không có cán bộ giảng dạy nào bị trùng lịch")
                return
        else:
            messagebox.showwarning(title="Chú ý",message="Không có thông tin lịch dạy vui lòng thêm lịch dạy")

    def bieudo(self):
        if self.cansave:
            x = []
            data_percent = []
            labels = []
            for i in self.list_:
                x.append(i._lec)

            unique_list = list(set(x))

            # for i in unique_list:
            #     data_percent.append(x.count(i))
            #     labels.append(i)

            for i in unique_list:
                point = 0
                for j in self.list_:
                    if i == j._lec:
                        point = point + int(j._credit)
                data_percent.append(point)
                labels.append(i)

            for i in range(0,len(labels)):
                if labels[i] == None:
                    labels[i] = 'Trống'
                    
            y = np.array(data_percent)

            explode = [0.1,0,0,0]
            plt.title("Biểu đồ khối lượng dạy của các giảng viên")
            plt.pie(y, labels=labels, autopct='%1.1f%%', wedgeprops={'edgecolor': 'white', 'linewidth': 1.5})
            plt.show()
        else:
            messagebox.showwarning(title="Chú ý",message="Không có thông tin lịch dạy vui lòng thêm lịch dạy")

    def taolichgg(self):
        if self.cansave:
            root = tk.Toplevel(self.window)
            ggcal.GoogleCalendar(root,self.list_)
        else:
            messagebox.showwarning(title="Chú ý",message="Không có thông tin lịch dạy vui lòng thêm lịch dạy")

    def creatfile():
        file_path = os.getcwd() + "\\alpha.xlsx"
        if not os.path.exists(file_path):
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "Sheet1"
            workbook.save(file_path)
            workbook.close()

    # Xếp giảng viên ưu tiên xếp hết
    def xepgiangvien(self):
        selected_table = self.status_combobox.get()
        if selected_table == "":
            messagebox.showwarning(title="Chú ý",message="Vui lòng chọn kỳ học")
            return
        try:
            self.treeview.delete(*self.treeview.get_children())
            if os.path.splitext(self.file_path)[1] == ".xlsx":
                self.list_xep = mn.readfile(self.file_path) 
                
            if os.path.splitext(self.file_path)[1] == ".xls":
                self.list_xep = mn.readfilexls(self.file_path)

            mn.checklopghep(self.list_xep)
            mn.check2lich(self.list_xep)
            mn.xepgiangvien(self.list_xep,selected_table)
            self.load_data(self.list_xep)
            self.list_ = self.list_xep
        except:
            messagebox.showwarning(title="Chú ý",message="Không có thông tin lịch dạy vui lòng thêm lịch dạy")
            return
        messagebox.showinfo(title="Thông báo",message="Sắp xếp cán bộ giảng dạy xong")

    # Xếp giảng viên ưu tiên xếp theo alpha
    def xepgiangvien1(self):
        selected_table = self.status_combobox.get()
        if selected_table == "":
            messagebox.showwarning(title="Chú ý",message="Vui lòng chọn kỳ học")
            return
        try:
            self.treeview.delete(*self.treeview.get_children())
            if os.path.splitext(self.file_path)[1] == ".xlsx":
                self.list_xep = mn.readfile(self.file_path) 
                
            if os.path.splitext(self.file_path)[1] == ".xls":
                self.list_xep = mn.readfilexls(self.file_path)

            mn.checklopghep(self.list_xep)
            mn.check2lich(self.list_xep)
            mn.xepgiangvien1(self.list_xep,selected_table)
            self.load_data(self.list_xep)
            self.list_ = self.list_xep
        except:
            messagebox.showwarning(title="Chú ý",message="Không có thông tin lịch dạy vui lòng thêm lịch dạy")
            return
        messagebox.showinfo(title="Thông báo",message="Sắp xếp cán bộ giảng dạy xong")

    def checkloptrung(self):
        if mn.canchecktrung(self.list_):
            self.treeview.delete(*self.treeview.get_children())
            mn.checktrung(self.list_)
            self.load_data(self.list_)
        else:
           messagebox.showwarning(title="Chú ý",message="Không có thông tin lịch dạy hoặc chưa xếp cán bộ giảng dạy vui lòng thêm thông tin") 
           return
        messagebox.showinfo(title="Thông báo",message="Kiểm tra trùng lịch xong")
        
    def save(self):
        if self.cansave:
            file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
            if file_path:
                # Tạo một workbook và một worksheet
                workbook = openpyxl.Workbook()
                worksheet = workbook.active
                worksheet.title = "Sheet1"
                date_style = NamedStyle(name='date_style')
                date_style.number_format = 'DD-MM-YYYY'
                
                head = ["STT","Mã học phần","Tên môn học","Mã lớp học","Lớp ghép","Thứ","Tiết","Phòng học","Số TC","Bắt đầu","Kết thúc","1234567890123456789012","Giảng viên"]
                worksheet.append(head)
                list_data = mn.list_save(self.list_)
                for i in list_data:
                    worksheet.append(i)
                # Đổi style ngày
                column_J = worksheet['J']
                for cell in column_J:
                    cell.style = date_style
                
                column_K = worksheet['K']
                for cell in column_K:
                    cell.style = date_style
                # Lưu workbook
                workbook.save(file_path)
                workbook.close()
                messagebox.showinfo(title="Thông báo",message="Lưu File thành công")
        else:
            messagebox.showwarning(title="Chú ý",message="Không có thông tin lịch dạy vui lòng thêm lịch dạy")
    
    def load_data(self,data):
        colors = ["#E6F1D8", "white"]
        row_index = 0
        # data dạng đối tượng rồi mới chuyển thành dạng list
        list_data = mn.list_print(data)

        for col_name in self.cols:
            self.treeview.heading(col_name, text=col_name)

        chuoi = list_data[0][1] 
        swap = ""
        for value in list_data:
            if chuoi == value[1]:
                bg_color = colors[0]
            else:
                swap = colors[0]
                colors[0] = colors[1]
                colors[1] = swap

                bg_color = colors[0]
                chuoi = value[1]
            value_list = list(value)
            for i in range(len(value_list)):
                if value_list[i] == None:
                    value_list[i] = '-'
            self.treeview.insert('', tk.END, values=value_list, tags=("my_font",bg_color))

            row_index = row_index + 1

        for color in colors:
            self.treeview.tag_configure(color, background=color)

    def rowfirst(self):
        selected_table = self.status_combobox.get()
        row = []
        workbook = openpyxl.load_workbook(self.path)
        sheet = workbook[selected_table]
        data = list(sheet.values)
        if len(data) == 0:
            return row
        else:
            row = data[0]
            return row[1:]
    
    def colfirst(self):
        selected_table = self.status_combobox.get()
        col = []
        workbook = openpyxl.load_workbook(self.path)
        sheet = workbook[selected_table]
        data = list(sheet.values)
        for i in data[1:]:
            col.append(i[0])
        return col
    
if __name__ == "__main__":
    # Guische.creatfile()
    window = Tk()
    Guische(window)
    window.mainloop()