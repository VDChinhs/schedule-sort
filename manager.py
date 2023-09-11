from openpyxl import Workbook, load_workbook
import xlrd
import lichhoc
import readalpha as ra
import math

list_ = []
listtest = []

# Tìm cột STT (vị trí)
def colSTT(excel):
    for i in range(1,excel.max_row): # type: ignore
        for j in range(1,excel.max_column):
            val = excel.cell(row = i, column = j).value # type: ignore
            if val == "STT":
                return j

# Tìm cột tiêu đề (vị trí)
def rowheading(excel):
    for i in range(1,excel.max_row): # type: ignore
        for j in range(1,excel.max_column):
            val = excel.cell(row = i, column = j).value # type: ignore
            if val == "STT":
                return i

def xeplopghep(sche,lich,somon,listlec,index):
    for i in sche:
        if lich._lopghep == i._stt and i._lec == None:
            i._lec = listlec[index]
            somon = somon + 1

def xeplop2lich(sche,lich,somon,listlec,index):
    for i in sche:
        if lich._course_name == i._course_name and lich._class_name == i._class_name and i._lec == None:
            i._lec = listlec[index]
            somon = somon + 1

def chuyengv(somon,maxx,index):
    if somon >= maxx:
        index = index + 1
        somon = 0

# Xếp giảng viên
def xepgiangvien(sche):
    listalpha = ra.listalpha()
    
    for tenmon in listalpha:
        index = 0
        somon = 0
        listlec = list(listalpha[tenmon].keys())
        listpoint = list(listalpha[tenmon].values())

        for lich in sche:
            if tenmon.lower().strip() == lich._course_name.lower().strip() and lich._lec == None and lich._course_name.strip().lower().count("đồ án") == 0:
                if len(listalpha[tenmon]) == 3:
                    # Tính tổng số môn mà gv được dạy
                    if index == listpoint.index(max(listpoint)):
                        maxx = math.floor(countmon(tenmon,sche)/3) + (countmon(tenmon,sche) % 3) + (int(listpoint[index]) - 3)
                    else:
                        maxx = math.floor(countmon(tenmon,sche)/3) + (int(listpoint[index]) - 3)

                    # Xét nếu mà có giảng viên ở lịch đó rồi không xếp nữa 
                    if lich._lec == None:
                        lich._lec = listlec[index]
                        somon = somon + 1

                        # Xếp gv cho lớp ghép
                        for i in sche:
                            if lich._lopghep == i._stt and i._lec == None:
                                i._lec = listlec[index]
                                somon = somon + 1
                                # Xếp gv cho lớp có 2 lịch
                                for j in sche:
                                    if j._course_name == i._course_name and j._class_name == i._class_name and j._lec == None:
                                        j._lec = listlec[index]
                                        somon = somon + 1

                        # Xếp gv cho lớp có 2 lịch dạy
                        for i in sche:
                            if lich._course_name == i._course_name and lich._class_name == i._class_name and i._lec == None:
                                i._lec = listlec[index]
                                somon = somon + 1

                        # Xếp gv cho lớp đồ án
                        for i in sche:
                            if i._course_name.strip().lower().count(lich._course_name.strip().lower()) > 0 and i._course_name.strip().lower().count("đồ án") != 0 and i._class_name == lich._class_name:
                                i._lec = listlec[index]

                    # Số môn gv dạy quá maxx thì chuyển xếp gv khác    
                    if somon >= maxx:
                        index = index + 1    
                        somon = 0

                if len(listalpha[tenmon]) == 2:
                    # Tính tổng số môn mà gv được dạy
                    if index == 0:
                        maxx = math.ceil(countmon(tenmon,sche)/2) + (int(listpoint[index]) - 5)
                    else:
                        maxx = math.floor(countmon(tenmon,sche)/2) + (int(listpoint[index]) - 5)
                    if maxx < 0:
                        maxx = 1

                    # Xét nếu mà có giảng viên ở lịch đó rồi không xếp nữa     
                    if lich._lec == None:
                        lich._lec = listlec[index]
                        somon = somon + 1
                        # Xếp gv cho lớp ghép
                        for i in sche:
                            if lich._lopghep == i._stt and i._lec == None:
                                i._lec = listlec[index]
                                somon = somon + 1
                                # Xếp gv cho lớp có 2 lịch
                                for j in sche:
                                        if j._course_name == i._course_name and j._class_name == i._class_name and j._lec == None:
                                            j._lec = listlec[index]
                                            somon = somon + 1
                            
                        # Xếp gv cho lớp có 2 lịch dạy
                        for i in sche:
                            if lich._course_name == i._course_name and lich._class_name == i._class_name and i._lec == None:
                                i._lec = listlec[index]
                                somon = somon + 1
                            
                        # Xếp gv cho những môn đồ án
                        for i in sche:
                            # Nếu có tên môn lại thêm chữ đồ án
                            if i._course_name.strip().lower().count(lich._course_name.strip().lower()) > 0 and i._course_name.strip().lower().count("đồ án") != 0 and i._class_name == lich._class_name and i._lec == None:
                                i._lec = listlec[index]

                    # Số môn gv dạy quá maxx thì chuyển xếp gv khác  
                    if somon >= maxx:
                        index = index + 1
                        somon = 0
                    
                if len(listalpha[tenmon]) == 1:
                    lich._lec = listlec[0]
                    for i in sche:
                        if i._course_name.strip().lower().count(lich._course_name.strip().lower()) > 0 and i._class_name == lich._class_name:
                            i._lec = listlec[0]
    
    # Xếp nốt nhưng môn đồ án chưa được xếp
    for i in sche:
        if i._lec == None and i._course_name.strip().lower().count("đồ án") != 0:
            for j in sche:
                if i != j and i._class_name == j._class_name and i._course_name.strip().lower().count(j._course_name.strip().lower()) != 0 and j._course_name.strip().lower().count("đồ án") == 0:
                    i._lec = j._lec

    # for tenmon in listalpha:
    #     for lich in sche:
    #         if tenmon.lower() == lich._course_name.lower():
    #             print(tenmon,lich._lec)
    #     print("____")


# Xếp lớp ghép
def lopghep1(sche):
    for i in range(len(sche)):
        if sche[i]._lopghep == None:
            for j in range(len(sche)):
                if sche[j]._lopghep == None:
                    # Cùng mã học phần, cùng thứ, cùng tiết
                    if(sche[i]._course_code == sche[j]._course_code and 
                    sche[i]._day == sche[j]._day and 
                    sche[i]._session == sche[j]._session and 
                    i != j):
                        sche[i]._lopghep = sche[j]._stt
                        sche[j]._lopghep = sche[i]._stt
                        break

def lopghep(sche):
    for i in sche:
        if i._lopghep == None:
            for j in sche:
                # Cùng mã học phần, cùng thứ, cùng tiết
                if j._lopghep == None:
                    if(i._course_code == j._course_code and i._day == j._day and i._session == j._session and i != j):
                        i._lopghep = j._stt
                        j._lopghep = i._stt
                        break

# Kiểm tra trùng lịch
def checktrung(sche):
    index = 1
    for i in range(len(sche)):
        cong = False
        for j in range(len(sche)):
            # Khác mã học phần, cùng thứ, cùng tiết, cùng giảng viên(các tuần có thể khác nhau)
            if(sche[i]._course_code != sche[j]._course_code and 
               sche[i]._day == sche[j]._day and 
               sche[i]._session == sche[j]._session and 
               sche[i]._lec == sche[j]._lec and
               i != j) and sche[i]._lopghep != sche[j]._stt and sche[j]._lopghep != sche[i]._stt and (sche[i]._lopghep == None and sche[j]._lopghep == None):
                # Còn trường hợp 3 lớp trùng nhau
                if isinstance(sche[i]._trung,int) and isinstance(sche[j]._trung,int) and sche[i]._trung == sche[j]._trung:
                    continue
                else:
                    #So sánh thứ tự tuần có thế lấy ra tuần trùng
                    for z in range(len("1234567890123456789012")):
                        if len(sche[i]._week[z]) < 22:
                            sche[i]._week = sche[i]._week + (22-len(sche[i]._week[z]))*" " 
                        if len(sche[j]._week[z]) < 22:
                            sche[j]._week = sche[j]._week + (22-len(sche[j]._week[z]))*" "     
                        if sche[i]._week[z] == sche[j]._week[z] and sche[i]._week[z] != " " and sche[j]._week[z] != " ":
                            # print(sche[i]._stt, sche[j]._stt,end="/")
                            sche[i]._trung = index
                            sche[j]._trung = index

                            sche[i]._tuantrung.append(sche[i]._week[z])
                            sche[j]._tuantrung.append(sche[j]._week[z])
                            # print(sche[i]._trung,sche[j]._trung,end="/")
                            # print(sche[i]._week[z])
                            cong = True
        if cong:            
            index = index + 1

# Tìm ngày bắt đầu năm học
def startday(sche):
    day = sche[0]._start
    for i in sche:
        if day > i._start:
            day = i._start
    return day

# Tìm ngày kết thúc năm học
def endday(sche):
    day = sche[0]._end
    for i in sche:
        if day < i._end:
            day = i._end
    return day

# Danh sách các môn học có tên là name (tạm thời không dùng đến nữa - thừa)
def countmon(name,sche):
    list_monhoc = []
    for i in sche:
        if i._course_name.lower() == name.lower():
            list_monhoc.append(i)
    return len(list_monhoc)


# Đọc File .xlxs (Hàm: truyền đường dẫn trả về list)
def readfile(path):
    ds = []
    wb = load_workbook(path)
    sh = wb.active
    for i in range(1,sh.max_row): # type: ignore
        val = sh.cell(row = i, column = colSTT(sh)).value # type: ignore
        # Kiểm tra có phải dòng cần đọc không
        if isinstance(val, int) or val is not None and val.isdigit():
            for j in range(1, sh.max_column + 1):
                # Đọc theo tên cột (File excel phải đúng tên cột)
                match sh.cell(row = rowheading(sh), column = j).value:
                    case "STT":
                        stt = sh.cell(row = i, column = j).value
                    case "Mã học phần":
                        cousre_code = sh.cell(row = i, column = j).value   
                    case "Tên môn học":
                        course_name = sh.cell(row = i, column = j).value
                    case "Mã lớp học":
                        class_name = sh.cell(row = i, column = j).value
                    case "Lớp ghép":
                        class_com = sh.cell(row = i, column = j).value
                    case "Thứ":
                        day = sh.cell(row = i, column = j).value
                    case "Tiết":
                        seesion = sh.cell(row = i, column = j).value
                    case "Phòng học":
                        room = sh.cell(row = i, column = j).value
                    case "Số TC":
                        credit = sh.cell(row = i, column = j).value
                    case "Bắt đầu":
                        start = sh.cell(row = i, column = j).value
                    case "Kết thúc":
                        end = sh.cell(row = i, column = j).value
                    case "1234567890123456789012":
                        week = sh.cell(row = i, column = j).value
                    case "Giảng viên":
                        lec = sh.cell(row = i, column = j).value
            mh = lichhoc.monhoc(stt,cousre_code,course_name,class_name,class_com,day,seesion,room,credit,start,end,week)
            ds.append(mh)
    wb.close()
    return ds      

# Chuyển lưu dạng đối tượng về dạng list
def list_print(sche):
    ds = []
    for i in sche:
        x = [
            i._stt,
            i._course_code,
            i._course_name,
            i._class_name,
            i._day,
            i._session,
            i._week,
            i._lec,
            i._lopghep,
            i._trung,
        ]
        ds.append(x)
    return ds

def list_save(sche):
    ds = []
    for i in sche:
        x = [
            i._stt,
            i._course_code,
            i._course_name,
            i._class_name,
            i._class_com,
            i._day,
            i._session,
            i._room,
            i._credit,
            i._start,
            i._end,
            i._week,
            i._lec,
        ]
        ds.append(x)
    return ds

list_ = readfile('tkb.xlsx')


# xepgiangvien(list_)
# checktrung(list_)

# listtest = list_print(list_)

# print(listtest)



# Read File .xls
# workbook = xlrd.open_workbook('D:\\Code Python in VSC\\phan_cong.xls')
# sheet = workbook["Sheet2"]

# for i in range(sheet.nrows):
#     val = sheet.cell_value(i, colSTT1(sheet))
#     if isinstance(val, int) or (val is not None and str(val).isdigit()):
#         for j in range(1,sheet.ncols):
#             match sheet.cell_value(rowheading1(sheet), j):
#                 case "STT":
#                     stt = sheet.cell_value(i, j)
#                 case "Mã học phần":
#                     cousre_code = sheet.cell_value(i, j)   
#                 case "Tên môn học":
#                     course_name = sheet.cell_value(i, j)
#                 case "Mã lớp học":
#                     class_name = sheet.cell_value(i, j)
#                 case "Lớp ghép":
#                     class_com = sheet.cell_value(i, j)
#                 case "Thứ":
#                     day = sheet.cell_value(i, j)
#                 case "Tiết":
#                     seesion = sheet.cell_value(i, j)
#                 case "Phòng học":
#                     room = sheet.cell_value(i, j)
#                 case "Số TC":
#                     credit = sheet.cell_value(i, j)
#                 case "Bắt đầu":
#                     start = sheet.cell_value(i, j)
#                 case "Kết thúc":
#                     end = sheet.cell_value(i, j)
#                 case "1234567890123456789012":
#                     week = sheet.cell_value(i, j)
#         mh = lichhoc.monhoc(stt,cousre_code,course_name,class_name,class_com,day,seesion,room,credit,start,end,week)
#         list.append(mh)
